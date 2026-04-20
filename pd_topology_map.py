#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PagerDuty Topology + Volume Exporter
- Services, Escalation Policies, Teams, Schedules
- Business Services (best-effort) and Service Dependencies (best-effort)
- Incident volume per service over a time window
- Exports multi-sheet XLSX + Edges, and generates DOT(+PNG), Mermaid, and Interactive HTML.

Requirements:
  pip install requests openpyxl
Optional:
  brew install graphviz   # for PNG from DOT

Usage:
  export PD_API_TOKEN="..."
  python pd_topology_map.py --volume-window last_30d --output xlsx --visual all --basename pd_topology_map --overwrite
"""

import argparse
import os
import sys
import time
import json
import shutil
import subprocess
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Tuple, Any, Optional

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


API_HOST = "https://api.pagerduty.com"
USER_AGENT = "pd-topology-exporter/1.0"
DEFAULT_PAGE_LIMIT = 100
REQUEST_TIMEOUT = 30  # seconds
MAX_RETRIES = 5


def iso_now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def iso_days_ago_utc(days: int) -> str:
    return (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")


def compute_window(window: str, start: Optional[str], end: Optional[str]) -> Tuple[str, str]:
    if window == "last_7d":
        return iso_days_ago_utc(7), iso_now_utc()
    if window == "last_30d":
        return iso_days_ago_utc(30), iso_now_utc()
    if window == "last_90d":
        return iso_days_ago_utc(90), iso_now_utc()
    if window == "custom":
        if not start or not end:
            raise ValueError("For custom window, both --start-iso and --end-iso are required.")
        # Basic validation
        for ts in (start, end):
            if "T" not in ts or not ts.endswith("Z"):
                raise ValueError("Timestamps must be ISO8601 in UTC, e.g., 2026-03-01T00:00:00Z")
        return start, end
    raise ValueError(f"Unknown window: {window}")


class PDClient:
    def __init__(self, token: str, host: str = API_HOST):
        self.host = host.rstrip("/")
        self.sess = requests.Session()
        self.sess.headers.update({
            "Authorization": f"Token token={token}",
            "Accept": "application/vnd.pagerduty+json;version=2",
            "Content-Type": "application/json",
            "User-Agent": USER_AGENT,
        })

    def _request(self, method: str, path: str, params: Optional[dict] = None, json_body: Optional[dict] = None) -> dict:
        url = f"{self.host}{path}"
        retries = 0
        while True:
            try:
                resp = self.sess.request(
                    method, url, params=params, json=json_body, timeout=REQUEST_TIMEOUT
                )
            except requests.RequestException as e:
                if retries < MAX_RETRIES:
                    sleep_s = 2 ** retries
                    print(f"Network error: {e}. Retrying in {sleep_s}s...", file=sys.stderr)
                    time.sleep(sleep_s)
                    retries += 1
                    continue
                raise

            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", "2"))
                print(f"Rate limited (429). Retrying in {retry_after}s...", file=sys.stderr)
                time.sleep(retry_after)
                retries += 1
                if retries > MAX_RETRIES:
                    resp.raise_for_status()
                continue

            if 500 <= resp.status_code < 600:
                if retries < MAX_RETRIES:
                    sleep_s = min(30, 2 ** retries)
                    print(f"Server error {resp.status_code}. Retrying in {sleep_s}s...", file=sys.stderr)
                    time.sleep(sleep_s)
                    retries += 1
                    continue

            if not resp.ok:
                try:
                    detail = resp.json()
                except Exception:
                    detail = resp.text
                raise requests.HTTPError(f"{method} {url} failed: {resp.status_code} {detail}")

            try:
                return resp.json()
            except ValueError:
                return {}

    def _paginate(self, path: str, root_key: str, params: Optional[dict] = None, include: Optional[List[str]] = None) -> List[dict]:
        params = dict(params or {})
        params.setdefault("limit", DEFAULT_PAGE_LIMIT)
        if include:
            # PD API supports repeated include[]=...
            # We'll add includes into params as list that requests encodes as multiple keys.
            params["include[]"] = include

        results = []
        offset = 0
        while True:
            params["offset"] = offset
            data = self._request("GET", path, params=params)
            batch = data.get(root_key, [])
            results.extend(batch)
            more = data.get("more", False)
            if not more:
                break
            offset = data.get("offset", 0) + data.get("limit", params["limit"])
        return results

    # Entity fetchers

    def get_services(self) -> List[dict]:
        # include escalation_policy and teams where supported
        return self._paginate("/services", "services", include=["escalation_policies", "teams"])

    def get_escalation_policies(self, ids: Optional[List[str]] = None) -> List[dict]:
        # Fetch EPs with all rules/targets
        params = {}
        if ids:
            # PD supports filter by ids[] for many endpoints
            # If not supported, we'll paginate all and filter client-side.
            # Try server-side first.
            params["ids[]"] = ids
        eps = self._paginate("/escalation_policies", "escalation_policies", params=params)
        if ids:
            # If API ignored filter, keep only the needed
            want = set(ids)
            eps = [e for e in eps if e.get("id") in want]
        return eps

    def get_schedules(self, ids: Optional[List[str]] = None) -> List[dict]:
        params = {}
        if ids:
            params["ids[]"] = ids
        sch = self._paginate("/schedules", "schedules", params=params)
        if ids:
            want = set(ids)
            sch = [s for s in sch if s.get("id") in want]
        return sch

    def get_teams(self) -> List[dict]:
        return self._paginate("/teams", "teams")

    def get_business_services(self) -> List[dict]:
        try:
            return self._paginate("/business_services", "business_services")
        except requests.HTTPError as e:
            print(f"Business services not available or endpoint error: {e}", file=sys.stderr)
            return []

    def get_service_dependencies_best_effort(self, service_ids: List[str]) -> List[dict]:
        """
        Best-effort fetch of service dependencies graph.
        Returns a list of relationship dicts:
          {
            "source_id": "...",
            "source_type": "business_service|service",
            "relation": "DEPENDS_ON|SUPPORTS",
            "target_id": "...",
            "target_type": "service|business_service"
          }
        Note: PagerDuty has a /service_dependencies endpoint, which may be behind features or vary.
        We'll try a few likely patterns and degrade gracefully.
        """
        edges: List[dict] = []

        # Attempt: GET /service_dependencies?ids[]=<service_id>
        # Aggregate minimally to avoid many calls; try batching up to, say, 50 ids per call.
        batch_size = 50
        tried = False
        for i in range(0, len(service_ids), batch_size):
            batch = service_ids[i:i + batch_size]
            params = [("ids[]", sid) for sid in batch]
            try:
                data = self._request("GET", "/service_dependencies", params=params)
                # Expecting something like {"relationships": [{"dependent_service": {...}, "supporting_service": {...}}, ...]}
                rels = data.get("relationships") or data.get("service_dependencies") or []
                for r in rels:
                    # Heuristic parsing
                    if "dependent_service" in r and "supporting_service" in r:
                        dep = r["dependent_service"]
                        sup = r["supporting_service"]
                        edges.append({
                            "source_id": dep.get("id"),
                            "source_type": "service",
                            "relation": "DEPENDS_ON",
                            "target_id": sup.get("id"),
                            "target_type": "service",
                        })
                    if "business_service" in r and "technical_service" in r:
                        bs = r["business_service"]
                        ts = r["technical_service"]
                        edges.append({
                            "source_id": bs.get("id"),
                            "source_type": "business_service",
                            "relation": "MAPPED_TO",
                            "target_id": ts.get("id"),
                            "target_type": "service",
                        })
                tried = True
            except requests.HTTPError:
                # Endpoint not available in this shape.
                break

        if not tried and service_ids:
            print("Service dependencies endpoint not available; skipping dependencies graph.", file=sys.stderr)
        return edges

    def count_incidents_for_service(self, service_id: str, since_iso: str, until_iso: str) -> int:
        """
        Counts incidents created for a given service in [since, until].
        Uses pagination on /incidents with service_ids[] filter.
        """
        total = 0
        offset = 0
        limit = 100
        params = {
            "since": since_iso,
            "until": until_iso,
            "service_ids[]": [service_id],
            "limit": limit,
            "offset": offset,
            "total": "true",
        }
        while True:
            params["offset"] = offset
            data = self._request("GET", "/incidents", params=params)
            batch = data.get("incidents", [])
            total += len(batch)
            more = data.get("more", False)
            if not more:
                break
            offset += limit
        return total


def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            max_length = max(max_length, len(v))
        ws.column_dimensions[col_letter].width = min(60, max(12, max_length + 2))


def write_sheet(wb: Workbook, title: str, headers: List[str], rows: List[List[Any]]):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    auto_fit_columns(ws)


def build_mermaid(nodes: Dict[str, dict], edges: List[dict]) -> str:
    # Mermaid graph TD with types grouped
    lines = ["graph TD"]
    # Declare nodes with type-based shapes (Mermaid doesn't have many shapes; we’ll use subgraph labels)
    # We'll just label nodes with [Type] Name
    for nid, n in nodes.items():
        label = n.get("name", nid).replace('"', "'")
        t = n.get("type", "node")
        lines.append(f'  {nid}["{label} ({t})"]')
    for e in edges:
        src = e["source_id"]
        dst = e["target_id"]
        rel = e["relation"]
        lines.append(f'  {src} -- "{rel}" --> {dst}')
    return "\n".join(lines)


def build_dot(nodes: Dict[str, dict], edges: List[dict]) -> str:
    # Graphviz DOT
    # Color/shape by type
    type_style = {
        "service": ('ellipse', 'lightblue'),
        "escalation_policy": ('box', 'gold'),
        "team": ('hexagon', 'palegreen'),
        "schedule": ('parallelogram', 'plum'),
        "business_service": ('doubleoctagon', 'orange'),
        "user": ('oval', 'white'),
    }
    lines = [
        "digraph PD {",
        '  graph [overlap=false, splines=true, fontsize=10];',
        '  node [style=filled, fontname="Helvetica", fontsize=10];',
        '  edge [fontname="Helvetica", fontsize=9];'
    ]
    for nid, n in nodes.items():
        t = n.get("type", "node")
        shape, color = type_style.get(t, ('ellipse', 'white'))
        label = n.get("name", nid).replace('"', "'")
        lines.append(f'  "{nid}" [label="{label}\\n({t})", shape={shape}, fillcolor="{color}"];')
    for e in edges:
        src = e["source_id"]
        dst = e["target_id"]
        rel = e["relation"]
        lines.append(f'  "{src}" -> "{dst}" [label="{rel}"];')
    lines.append("}")
    return "\n".join(lines)


def build_interactive_html(nodes: Dict[str, dict], edges: List[dict]) -> str:
    # Self-contained HTML using vis-network from CDN
    node_items = []
    for nid, n in nodes.items():
        node_items.append({
            "id": nid,
            "label": f"{n.get('name', nid)}\n({n.get('type')})",
            "group": n.get("type"),
            "title": json.dumps(n, ensure_ascii=False),
        })
    edge_items = []
    for e in edges:
        edge_items.append({
            "from": e["source_id"],
            "to": e["target_id"],
            "label": e["relation"],
            "arrows": "to"
        })
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<title>PagerDuty Topology</title>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<link rel="preconnect" href="https://cdnjs.cloudflare.com">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/vis-network/9.1.9/dist/vis-network.min.css">
<style>
  html, body {{ height: 100%; margin: 0; }}
  #network {{ width: 100%; height: 98vh; border: 1px solid #ddd; }}
  .legend {{ padding: 8px; font-family: Arial; font-size: 12px; }}
  .legend span {{ display: inline-block; margin-right: 10px; }}
</style>
</head>
<body>
<div class="legend">
  <strong>Legend:</strong>
  <span>Service</span>
  <span>Escalation Policy</span>
  <span>Team</span>
  <span>Schedule</span>
  <span>Business Service</span>
</div>
<div id="network"></div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/vis-network/9.1.9/dist/vis-network.min.js"></script>
<script>
const nodes = new vis.DataSet({json.dumps(node_items)});
const edges = new vis.DataSet({json.dumps(edge_items)});
const container = document.getElementById('network');
const data = {{ nodes, edges }};
const options = {{
  interaction: {{ hover: true }},
  nodes: {{
    shape: 'dot',
    scaling: {{ min: 8, max: 30 }}
  }},
  groups: {{
    service: {{ color: '#ADD8E6' }},
    escalation_policy: {{ color: '#FFD700' }},
    team: {{ color: '#98FB98' }},
    schedule: {{ color: '#DDA0DD' }},
    business_service: {{ color: '#FFA500' }},
    user: {{ color: '#FFFFFF' }}
  }},
  physics: {{
    stabilization: {{ iterations: 150 }},
    solver: 'barnesHut'
  }}
}};
const network = new vis.Network(container, data, options);
</script>
</body>
</html>"""
    return html


def main():
    parser = argparse.ArgumentParser(description="PagerDuty topology and volume exporter")
    parser.add_argument("--api-host", default=API_HOST, help="API host base URL (default: US region)")
    parser.add_argument("--volume-window", choices=["last_7d", "last_30d", "last_90d", "custom"], default="last_30d")
    parser.add_argument("--start-iso", help="Custom start ISO8601 (UTC, ends with Z) if --volume-window=custom")
    parser.add_argument("--end-iso", help="Custom end ISO8601 (UTC, ends with Z) if --volume-window=custom")
    parser.add_argument("--output", choices=["xlsx"], default="xlsx", help="Output format for tables (xlsx)")
    parser.add_argument("--visual", choices=["all", "dot", "mermaid", "html"], default="all", help="Visual outputs to generate")
    parser.add_argument("--basename", default="pd_topology_map", help="Base filename for outputs")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing files")
    parser.add_argument("--include-users", action="store_true", help="Include Users and OnCall (you set this to false)")
    parser.add_argument("--include-business-services", action="store_true", help="Include business services and dependencies")
    args = parser.parse_args()

    token = os.environ.get("PD_API_TOKEN")
    if not token:
        print("ERROR: Please set PD_API_TOKEN environment variable.", file=sys.stderr)
        sys.exit(1)

    # Window
    since_iso, until_iso = compute_window(args.volume_window, args.start_iso, args.end_iso)

    # File existence checks
    out_files = []
    if args.output == "xlsx":
        out_files.append(f"{args.basename}.xlsx")
    if args.visual in ("all", "dot"):
        out_files.append(f"{args.basename}.dot")
        out_files.append(f"{args.basename}.png")  # may or may not be produced; check later
    if args.visual in ("all", "mermaid"):
        out_files.append(f"{args.basename}.mermaid.md")
    if args.visual in ("all", "html"):
        out_files.append(f"{args.basename}.html")

    if not args.overwrite:
        clashes = [f for f in out_files if os.path.exists(f)]
        if clashes:
            print(f"Refusing to overwrite existing files without --overwrite: {', '.join(clashes)}", file=sys.stderr)
            sys.exit(2)

    client = PDClient(token=token, host=args.api_host)

    print("Fetching services...")
    services = client.get_services()

    # Build basic nodes
    nodes: Dict[str, dict] = {}
    edges: List[dict] = []

    # Collect EP and schedule IDs
    ep_ids: set = set()
    schedule_ids: set = set()
    team_ids: set = set()

    for s in services:
        sid = s.get("id")
        nodes[sid] = {
            "id": sid,
            "type": "service",
            "name": s.get("name") or sid,
            "html_url": s.get("html_url"),
            "status": s.get("status"),
            "description": (s.get("description") or "")[:3000],
        }
        ep = s.get("escalation_policy")
        if ep and isinstance(ep, dict):
            ep_id = ep.get("id")
            if ep_id:
                ep_ids.add(ep_id)
                nodes.setdefault(ep_id, {"id": ep_id, "type": "escalation_policy", "name": ep.get("name", ep_id)})
                edges.append({
                    "source_id": sid, "source_type": "service",
                    "relation": "USES",
                    "target_id": ep_id, "target_type": "escalation_policy"
                })

        # Teams may be included
        tlist = s.get("teams") or []
        for t in tlist:
            tid = t.get("id")
            if not tid:
                continue
            team_ids.add(tid)
            nodes.setdefault(tid, {"id": tid, "type": "team", "name": t.get("name", tid)})
            edges.append({
                "source_id": sid, "source_type": "service",
                "relation": "OWNED_BY",
                "target_id": tid, "target_type": "team"
            })

    # Fetch full EPs to get schedules
    if ep_ids:
        print(f"Fetching escalation policies ({len(ep_ids)})...")
        eps = client.get_escalation_policies(list(ep_ids))
    else:
        eps = []

    ep_by_id = {e.get("id"): e for e in eps}
    for ep in eps:
        eid = ep.get("id")
        nodes.setdefault(eid, {"id": eid, "type": "escalation_policy", "name": ep.get("name", eid)})
        rules = ep.get("escalation_rules") or []
        for r in rules:
            targets = r.get("targets") or []
            for t in targets:
                if t.get("type", "").startswith("schedule"):
                    sch = t
                    sch_id = sch.get("id")
                    if sch_id:
                        schedule_ids.add(sch_id)
                        # node later after fetch
                        edges.append({
                            "source_id": eid, "source_type": "escalation_policy",
                            "relation": "REFERENCES",
                            "target_id": sch_id, "target_type": "schedule"
                        })

    # Fetch schedules
    schedules = []
    if schedule_ids:
        print(f"Fetching schedules ({len(schedule_ids)})...")
        schedules = client.get_schedules(list(schedule_ids))
        for sch in schedules:
            sid = sch.get("id")
            if not sid:
                continue
            nodes.setdefault(sid, {"id": sid, "type": "schedule", "name": sch.get("name", sid), "time_zone": sch.get("time_zone")})

    # Teams catalog (ensure names are populated)
    print("Fetching teams catalog...")
    teams = client.get_teams()
    for t in teams:
        tid = t.get("id")
        if tid and tid in nodes and nodes[tid].get("type") == "team":
            nodes[tid]["name"] = t.get("name", tid)

    # Business services and dependencies (best-effort)
    business_services = []
    dep_edges = []
    if args.include_business_services:
        print("Fetching business services (best-effort)...")
        business_services = client.get_business_services()
        for bs in business_services:
            bid = bs.get("id")
            if not bid:
                continue
            nodes.setdefault(bid, {"id": bid, "type": "business_service", "name": bs.get("name", bid)})

        # Try dependencies graph
        dep_edges = client.get_service_dependencies_best_effort([s.get("id") for s in services if s.get("id")])
        for e in dep_edges:
            # Ensure nodes exist
            sid = e["source_id"]
            tid = e["target_id"]
            if sid not in nodes:
                nodes[sid] = {"id": sid, "type": e.get("source_type", "service"), "name": sid}
            if tid not in nodes:
                nodes[tid] = {"id": tid, "type": e.get("target_type", "service"), "name": tid}
        edges.extend(dep_edges)

    # Incident volume per service
    print("Counting incidents per service over window...")
    since_iso, until_iso = compute_window(args.volume_window, args.start_iso, args.end_iso)
    incident_counts: Dict[str, int] = {}
    for s in services:
        sid = s.get("id")
        if not sid:
            continue
        try:
            cnt = client.count_incidents_for_service(sid, since_iso, until_iso)
        except Exception as e:
            print(f"Warning: failed to count incidents for service {sid}: {e}", file=sys.stderr)
            cnt = 0
        incident_counts[sid] = cnt
        # annotate node
        if sid in nodes:
            nodes[sid]["incident_count"] = cnt
        # progress feedback, lightweight
        if len(incident_counts) % 20 == 0:
            print(f"  processed {len(incident_counts)}/{len(services)} services...")

    # Build tabular data
    print("Building tabular export...")

    services_rows = []
    for s in services:
        sid = s.get("id")
        ep = s.get("escalation_policy") or {}
        ep_id = ep.get("id") if isinstance(ep, dict) else None
        ep_name = ep.get("name") if isinstance(ep, dict) else None
        team_names = ", ".join([t.get("name", t.get("id", "")) for t in (s.get("teams") or [])])
        services_rows.append([
            sid,
            s.get("name"),
            s.get("status"),
            s.get("html_url"),
            ep_id,
            ep_name,
            team_names,
            incident_counts.get(sid, 0),
            s.get("description") or ""
        ])

    ep_rows = []
    for eid, ep in ep_by_id.items():
        rules = ep.get("escalation_rules") or []
        ep_rows.append([eid, ep.get("name"), len(rules), ep.get("html_url")])

    team_rows = []
    for t in teams:
        team_rows.append([t.get("id"), t.get("name"), t.get("html_url")])

    schedule_rows = []
    for sch in schedules:
        schedule_rows.append([sch.get("id"), sch.get("name"), sch.get("time_zone"), sch.get("html_url")])

    bs_rows = []
    for bs in business_services:
        bs_rows.append([bs.get("id"), bs.get("name"), bs.get("html_url")])

    edge_rows = []
    for e in edges:
        edge_rows.append([
            e.get("source_id"),
            e.get("source_type"),
            e.get("relation"),
            e.get("target_id"),
            e.get("target_type"),
        ])

    # Write XLSX
    if args.output == "xlsx":
        xlsx_path = f"{args.basename}.xlsx"
        if os.path.exists(xlsx_path) and args.overwrite:
            os.remove(xlsx_path)
        wb = Workbook()
        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        write_sheet(
            wb,
            "Services",
            ["service_id", "name", "status", "html_url", "escalation_policy_id", "escalation_policy_name", "teams", "incident_count", "description"],
            services_rows,
        )
        write_sheet(
            wb,
            "EscalationPolicies",
            ["ep_id", "name", "num_rules", "html_url"],
            ep_rows,
        )
        write_sheet(
            wb,
            "Teams",
            ["team_id", "name", "html_url"],
            team_rows,
        )
        write_sheet(
            wb,
            "Schedules",
            ["schedule_id", "name", "time_zone", "html_url"],
            schedule_rows,
        )
        if args.include_business_services:
            write_sheet(
                wb,
                "BusinessServices",
                ["business_service_id", "name", "html_url"],
                bs_rows,
            )
        write_sheet(
            wb,
            "Edges",
            ["source_id", "source_type", "relation", "target_id", "target_type"],
            edge_rows,
        )
        wb.save(xlsx_path)
        print(f"Wrote {xlsx_path}")

    # Build visuals
    # Prepare a slimmed node dictionary for labels
    vis_nodes: Dict[str, dict] = {}
    for nid, n in nodes.items():
        label = n.get("name") or nid
        if n.get("type") == "service" and "incident_count" in n:
            label = f"{label} [{n['incident_count']}]"
        vis_nodes[nid] = {
            "id": nid,
            "type": n.get("type"),
            "name": label
        }

    if args.visual in ("all", "mermaid"):
        mermaid_txt = build_mermaid(vis_nodes, edges)
        mermaid_path = f"{args.basename}.mermaid.md"
        if os.path.exists(mermaid_path) and args.overwrite:
            os.remove(mermaid_path)
        with open(mermaid_path, "w", encoding="utf-8") as f:
            f.write("```mermaid\n")
            f.write(mermaid_txt)
            f.write("\n```")
        print(f"Wrote {mermaid_path}")

    if args.visual in ("all", "dot"):
        dot_txt = build_dot(vis_nodes, edges)
        dot_path = f"{args.basename}.dot"
        if os.path.exists(dot_path) and args.overwrite:
            os.remove(dot_path)
        with open(dot_path, "w", encoding="utf-8") as f:
            f.write(dot_txt)
        print(f"Wrote {dot_path}")

        # Try to render PNG if 'dot' exists
        dot_bin = shutil.which("dot")
        png_path = f"{args.basename}.png"
        if dot_bin:
            try:
                if os.path.exists(png_path) and args.overwrite:
                    os.remove(png_path)
                subprocess.check_call([dot_bin, "-Tpng", dot_path, "-o", png_path])
                print(f"Wrote {png_path}")
            except subprocess.CalledProcessError as e:
                print(f"Graphviz 'dot' failed to render PNG: {e}. Kept DOT file.", file=sys.stderr)
        else:
            print("Graphviz 'dot' not found. Skipping PNG render. Install via 'brew install graphviz' to enable.", file=sys.stderr)

    if args.visual in ("all", "html"):
        html_txt = build_interactive_html(vis_nodes, edges)
        html_path = f"{args.basename}.html"
        if os.path.exists(html_path) and args.overwrite:
            os.remove(html_path)
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_txt)
        print(f"Wrote {html_path}")

    print("Done.")


if __name__ == "__main__":
    main()
